import io
import re

from telegram import Update
from telegram.ext import (
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)
from openpyxl import Workbook, load_workbook

import database as db
import keyboards as kb
from config import ADMIN_IDS, ORDER_STATUSES

# مراحل مکالمه افزودن محصول
ADD_CAT_NAME, ADD_PROD_CAT, ADD_PROD_NAME, ADD_PROD_UNIT, ADD_PROD_PRICE = range(5)
# مراحل تغییر قیمت قدیمی
EDIT_PRICE_PICK, EDIT_PRICE_VALUE = range(5, 7)
# مرحله ایمپورت اکسل
AWAIT_EXCEL_FILE = 7
# مراحل ویرایش کامل محصول
MANAGE_EDIT_CATEGORY, MANAGE_EDIT_NAME, MANAGE_EDIT_UNIT, MANAGE_EDIT_PRICE = range(8, 12)


PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
HEADER_ALIASES = {
    "category": {
        "دسته بندی", "دسته‌بندی", "دسته", "گروه", "گروه کالا", "category", "category name"
    },
    "name": {
        "نام محصول", "محصول", "نام کالا", "کالا", "شرح کالا", "product", "product name"
    },
    "unit": {"واحد", "واحد اندازه گیری", "واحد اندازه‌گیری", "unit"},
    "price": {"قیمت", "قیمت فروش", "فی", "نرخ", "price", "sale price"},
    "active": {"وضعیت", "فعال", "فعال بودن", "status", "is active"},
}


def is_admin(user_id):
    return user_id in ADMIN_IDS


async def admin_only_guard(update: Update):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ این بخش فقط برای ادمین‌هاست.")
        return False
    return True


async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_only_guard(update):
        return
    await update.message.reply_text("پنل مدیریت:", reply_markup=kb.admin_menu_keyboard())


async def back_to_user_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("بازگشت به منوی اصلی.", reply_markup=kb.main_menu_keyboard())


async def cancel_admin_operation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    for key in (
        "new_prod_cat_id", "new_prod_name", "new_prod_unit", "editing_product_id",
        "manage_edit_product_id", "manage_edit_page", "manage_edit_category_id",
        "manage_edit_name", "manage_edit_unit",
    ):
        context.user_data.pop(key, None)
    await update.message.reply_text("عملیات لغو شد.", reply_markup=kb.admin_menu_keyboard())
    return ConversationHandler.END


# ---------- افزودن محصول ----------

async def add_product_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_only_guard(update):
        return ConversationHandler.END
    cats = db.list_categories()
    if cats:
        lines = "\n".join(f"- {c['name']}" for c in cats[:30])
        await update.message.reply_text(
            f"دسته‌بندی‌های موجود:\n{lines}\n\n"
            "نام دسته‌بندی محصول جدید را بفرستید. اگر جدید باشد، خودکار ساخته می‌شود:"
        )
    else:
        await update.message.reply_text("نام دسته‌بندی محصول را بفرستید؛ مثلاً میلگرد:")
    return ADD_CAT_NAME


async def add_product_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cat_name = update.message.text.strip()
    try:
        cat_id = db.add_category(cat_name)
    except ValueError:
        await update.message.reply_text("نام دسته‌بندی نمی‌تواند خالی باشد.")
        return ADD_CAT_NAME
    context.user_data["new_prod_cat_id"] = cat_id
    await update.message.reply_text("نام محصول را بفرستید؛ مثلاً میلگرد ۱۴ آجدار:")
    return ADD_PROD_NAME


async def add_product_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("نام محصول نمی‌تواند خالی باشد.")
        return ADD_PROD_NAME
    context.user_data["new_prod_name"] = name
    await update.message.reply_text("واحد اندازه‌گیری چیست؟ مثلاً کیلوگرم، تن یا شاخه:")
    return ADD_PROD_UNIT


async def add_product_unit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    unit = update.message.text.strip()
    if not unit:
        await update.message.reply_text("واحد نمی‌تواند خالی باشد.")
        return ADD_PROD_UNIT
    context.user_data["new_prod_unit"] = unit
    await update.message.reply_text("قیمت به تومان را به عدد بفرستید؛ مثلاً 285000:")
    return ADD_PROD_PRICE


def parse_price(value):
    if value is None or isinstance(value, bool):
        raise ValueError("قیمت خالی یا نامعتبر است.")
    if isinstance(value, (int, float)):
        if value < 0:
            raise ValueError("قیمت منفی است.")
        return int(round(value))

    text = str(value).translate(PERSIAN_DIGITS).strip().lower()
    text = text.replace("تومان", "").replace("ریال", "")
    text = re.sub(r"[,،٬\s]", "", text)
    if not text:
        raise ValueError("قیمت خالی است.")
    try:
        number = float(text)
    except ValueError as exc:
        raise ValueError("قیمت عددی نیست.") from exc
    if number < 0:
        raise ValueError("قیمت منفی است.")
    return int(round(number))


async def add_product_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        price = parse_price(update.message.text)
    except ValueError:
        await update.message.reply_text("لطفاً فقط قیمت معتبر به عدد بفرستید.")
        return ADD_PROD_PRICE

    cat_id = context.user_data["new_prod_cat_id"]
    name = context.user_data["new_prod_name"]
    unit = context.user_data["new_prod_unit"]
    product_id = db.add_product(cat_id, name, unit, price)
    await update.message.reply_text(
        f"✅ محصول «{name}» با قیمت {price:,} تومان / {unit} اضافه شد. (کد: {product_id})",
        reply_markup=kb.admin_menu_keyboard(),
    )
    return ConversationHandler.END


# ---------- مدیریت کامل محصولات ----------

def product_detail_text(product):
    category = db.get_category(product["category_id"])
    status = "🟢 فعال" if product["is_active"] else "⚪ غیرفعال"
    return (
        f"🧰 مدیریت محصول #{product['id']}\n\n"
        f"نام: {product['name']}\n"
        f"دسته‌بندی: {category['name'] if category else '---'}\n"
        f"واحد: {product['unit']}\n"
        f"قیمت: {product['price']:,} تومان\n"
        f"وضعیت: {status}\n\n"
        "عملیات موردنظر را انتخاب کنید:"
    )


async def manage_products_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_only_guard(update):
        return
    products = db.list_products(active_only=False)
    if not products:
        await update.message.reply_text(
            "هنوز محصولی ثبت نشده است.", reply_markup=kb.admin_menu_keyboard()
        )
        return
    active_count = sum(1 for p in products if p["is_active"])
    await update.message.reply_text(
        f"🧰 مدیریت محصولات\nتعداد کل: {len(products)} | فعال: {active_count} | غیرفعال: {len(products) - active_count}\n\n"
        "برای ویرایش، فعال/غیرفعال‌کردن یا حذف، یک محصول را انتخاب کنید:",
        reply_markup=kb.product_management_keyboard(products, page=0),
    )


async def manage_products_page(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.answer("فقط ادمین اجازه دسترسی دارد.", show_alert=True)
        return
    page = int(query.data.split(":")[1])
    products = db.list_products(active_only=False)
    if not products:
        await query.edit_message_text("محصولی برای مدیریت وجود ندارد.")
        return
    active_count = sum(1 for p in products if p["is_active"])
    await query.edit_message_text(
        f"🧰 مدیریت محصولات\nتعداد کل: {len(products)} | فعال: {active_count} | غیرفعال: {len(products) - active_count}\n\n"
        "یک محصول را انتخاب کنید:",
        reply_markup=kb.product_management_keyboard(products, page=page),
    )


async def manage_product_open(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.answer("فقط ادمین اجازه دسترسی دارد.", show_alert=True)
        return
    _, product_id_text, page_text = query.data.split(":")
    product = db.get_product(int(product_id_text))
    if not product:
        await query.edit_message_text("این محصول دیگر وجود ندارد.")
        return
    await query.edit_message_text(
        product_detail_text(product),
        reply_markup=kb.product_actions_keyboard(
            product["id"], bool(product["is_active"]), int(page_text)
        ),
    )


async def manage_product_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.answer("فقط ادمین اجازه این کار را دارد.", show_alert=True)
        return
    _, product_id_text, page_text = query.data.split(":")
    product_id = int(product_id_text)
    product = db.get_product(product_id)
    if not product:
        await query.edit_message_text("این محصول دیگر وجود ندارد.")
        return
    db.set_product_active(product_id, not bool(product["is_active"]))
    product = db.get_product(product_id)
    await query.edit_message_text(
        product_detail_text(product),
        reply_markup=kb.product_actions_keyboard(
            product_id, bool(product["is_active"]), int(page_text)
        ),
    )


async def manage_product_delete_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.answer("فقط ادمین اجازه این کار را دارد.", show_alert=True)
        return
    _, product_id_text, page_text = query.data.split(":")
    product = db.get_product(int(product_id_text))
    if not product:
        await query.edit_message_text("این محصول دیگر وجود ندارد.")
        return
    await query.edit_message_text(
        f"⚠️ آیا محصول «{product['name']}» برای همیشه حذف شود؟\n"
        "این کار قابل بازگشت نیست. برای مخفی‌کردن موقت، بهتر است محصول را غیرفعال کنید.",
        reply_markup=kb.product_delete_confirm_keyboard(product["id"], int(page_text)),
    )


async def manage_product_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.answer("فقط ادمین اجازه این کار را دارد.", show_alert=True)
        return
    _, product_id_text, page_text = query.data.split(":")
    product_id = int(product_id_text)
    product = db.get_product(product_id)
    if not product:
        await query.edit_message_text("این محصول قبلاً حذف شده است.")
        return
    name = product["name"]
    db.delete_product(product_id)
    products = db.list_products(active_only=False)
    if not products:
        await query.edit_message_text(f"✅ محصول «{name}» حذف شد. اکنون محصول دیگری وجود ندارد.")
        return
    page = int(page_text)
    max_page = max(0, (len(products) - 1) // 8)
    page = min(page, max_page)
    await query.edit_message_text(
        f"✅ محصول «{name}» حذف شد.\n\nیک محصول دیگر را انتخاب کنید:",
        reply_markup=kb.product_management_keyboard(products, page=page),
    )


async def manage_close(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        return
    await query.edit_message_text("مدیریت محصولات بسته شد. از منوی پایین گزینه بعدی را انتخاب کنید.")


async def manage_noop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()


async def manage_product_edit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.answer("فقط ادمین اجازه این کار را دارد.", show_alert=True)
        return ConversationHandler.END
    _, product_id_text, page_text = query.data.split(":")
    product = db.get_product(int(product_id_text))
    if not product:
        await query.edit_message_text("این محصول دیگر وجود ندارد.")
        return ConversationHandler.END

    category = db.get_category(product["category_id"])
    context.user_data["manage_edit_product_id"] = product["id"]
    context.user_data["manage_edit_page"] = int(page_text)
    context.user_data["manage_edit_category_id"] = product["category_id"]
    context.user_data["manage_edit_name"] = product["name"]
    context.user_data["manage_edit_unit"] = product["unit"]

    await query.edit_message_text(
        f"✏️ ویرایش «{product['name']}»\n\n"
        f"دسته فعلی: {category['name'] if category else '---'}\n"
        "نام دسته‌بندی جدید را بفرستید؛ برای نگه‌داشتن مقدار فعلی فقط - بفرستید.\n"
        "برای لغو کامل /cancel را بزنید."
    )
    return MANAGE_EDIT_CATEGORY


async def manage_edit_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text != "-":
        try:
            context.user_data["manage_edit_category_id"] = db.add_category(text)
        except ValueError:
            await update.message.reply_text("نام دسته‌بندی معتبر نیست؛ دوباره بفرستید.")
            return MANAGE_EDIT_CATEGORY
    current = context.user_data["manage_edit_name"]
    await update.message.reply_text(
        f"نام فعلی: {current}\nنام جدید را بفرستید؛ برای نگه‌داشتن مقدار فعلی - بفرستید:"
    )
    return MANAGE_EDIT_NAME


async def manage_edit_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text != "-":
        if not text:
            await update.message.reply_text("نام محصول نمی‌تواند خالی باشد.")
            return MANAGE_EDIT_NAME
        context.user_data["manage_edit_name"] = text
    current = context.user_data["manage_edit_unit"]
    await update.message.reply_text(
        f"واحد فعلی: {current}\nواحد جدید را بفرستید؛ برای نگه‌داشتن مقدار فعلی - بفرستید:"
    )
    return MANAGE_EDIT_UNIT


async def manage_edit_unit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text != "-":
        if not text:
            await update.message.reply_text("واحد نمی‌تواند خالی باشد.")
            return MANAGE_EDIT_UNIT
        context.user_data["manage_edit_unit"] = text
    product = db.get_product(context.user_data["manage_edit_product_id"])
    if not product:
        await update.message.reply_text("محصول دیگر وجود ندارد.", reply_markup=kb.admin_menu_keyboard())
        return ConversationHandler.END
    await update.message.reply_text(
        f"قیمت فعلی: {product['price']:,} تومان\n"
        "قیمت جدید را بفرستید؛ برای نگه‌داشتن مقدار فعلی - بفرستید:"
    )
    return MANAGE_EDIT_PRICE


async def manage_edit_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    product_id = context.user_data.get("manage_edit_product_id")
    product = db.get_product(product_id) if product_id else None
    if not product:
        await update.message.reply_text("محصول دیگر وجود ندارد.", reply_markup=kb.admin_menu_keyboard())
        return ConversationHandler.END

    text = update.message.text.strip()
    if text == "-":
        price = product["price"]
    else:
        try:
            price = parse_price(text)
        except ValueError:
            await update.message.reply_text("قیمت معتبر نیست. عدد بفرستید یا برای حفظ قیمت فعلی - بفرستید.")
            return MANAGE_EDIT_PRICE

    db.update_product(
        product_id,
        context.user_data["manage_edit_category_id"],
        context.user_data["manage_edit_name"],
        context.user_data["manage_edit_unit"],
        price,
    )
    updated = db.get_product(product_id)
    page = context.user_data.get("manage_edit_page", 0)
    await update.message.reply_text(
        "✅ اطلاعات محصول با موفقیت ویرایش شد.\n\n" + product_detail_text(updated),
        reply_markup=kb.product_actions_keyboard(
            updated["id"], bool(updated["is_active"]), page
        ),
    )
    for key in (
        "manage_edit_product_id", "manage_edit_page", "manage_edit_category_id",
        "manage_edit_name", "manage_edit_unit",
    ):
        context.user_data.pop(key, None)
    return ConversationHandler.END


# ---------- تغییر قیمت قدیمی (برای سازگاری) ----------

async def edit_price_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_only_guard(update):
        return ConversationHandler.END
    cats = db.list_categories()
    if not cats:
        await update.message.reply_text("هنوز محصولی ثبت نشده.")
        return ConversationHandler.END
    await update.message.reply_text(
        "یک دسته‌بندی انتخاب کنید:", reply_markup=kb.categories_keyboard(prefix="editcat")
    )
    return ConversationHandler.END


async def edit_price_pick_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.edit_message_text("⛔ فقط برای ادمین.")
        return
    cat_id = int(query.data.split(":")[1])
    products = db.list_products(category_id=cat_id, active_only=False)
    if not products:
        await query.edit_message_text("محصولی در این دسته نیست.")
        return
    await query.edit_message_text(
        "محصول موردنظر برای تغییر قیمت را انتخاب کنید:",
        reply_markup=kb.products_keyboard(cat_id, prefix="editprod", active_only=False),
    )


async def edit_price_pick_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.edit_message_text("⛔ فقط برای ادمین.")
        return ConversationHandler.END
    product_id = int(query.data.split(":")[1])
    context.user_data["editing_product_id"] = product_id
    product = db.get_product(product_id)
    if not product:
        await query.edit_message_text("محصول پیدا نشد.")
        return ConversationHandler.END
    await query.edit_message_text(
        f"قیمت فعلی «{product['name']}»: {product['price']:,} تومان / {product['unit']}\n\n"
        "قیمت جدید را به عدد بفرستید:"
    )
    return EDIT_PRICE_VALUE


async def edit_price_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        new_price = parse_price(update.message.text)
    except ValueError:
        await update.message.reply_text("لطفاً فقط عدد معتبر بفرستید.")
        return EDIT_PRICE_VALUE
    product_id = context.user_data["editing_product_id"]
    db.update_price(product_id, new_price)
    product = db.get_product(product_id)
    await update.message.reply_text(
        f"✅ قیمت «{product['name']}» به {new_price:,} تومان به‌روز شد.",
        reply_markup=kb.admin_menu_keyboard(),
    )
    return ConversationHandler.END


# ---------- ورود/خروج اکسل ----------

def normalize_header(value):
    if value is None:
        return ""
    text = str(value).translate(PERSIAN_DIGITS).lower()
    text = text.replace("ي", "ی").replace("ك", "ک").replace("‌", " ")
    text = re.sub(r"[_\-–—|:/\\]+", " ", text)
    return " ".join(text.strip().split())


def detect_header_map(sheet):
    normalized_aliases = {
        field: {normalize_header(alias) for alias in aliases}
        for field, aliases in HEADER_ALIASES.items()
    }
    max_scan_row = min(max(sheet.max_row, 1), 10)
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan_row, values_only=True), start=1
    ):
        mapping = {}
        for index, value in enumerate(row):
            header = normalize_header(value)
            for field, aliases in normalized_aliases.items():
                if header in aliases and field not in mapping:
                    mapping[field] = index
                    break
        if {"category", "name", "unit", "price"}.issubset(mapping):
            return row_number, mapping

    # سازگاری با قالب قدیمی A تا D، در صورتی که عنوان ستون‌ها متفاوت باشد.
    if sheet.max_column >= 4:
        return 1, {"category": 0, "name": 1, "unit": 2, "price": 3}
    return None, None


def cell_value(row, index):
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_active(value):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        return value
    text = normalize_header(value).translate(PERSIAN_DIGITS)
    if text in {"1", "فعال", "بله", "yes", "true", "active"}:
        return True
    if text in {"0", "غیرفعال", "غیر فعال", "خیر", "no", "false", "inactive"}:
        return False
    raise ValueError("وضعیت باید فعال یا غیرفعال باشد.")


async def import_excel_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_only_guard(update):
        return ConversationHandler.END
    await update.message.reply_text(
        "یک فایل اکسل (.xlsx) بفرستید. ستون‌های الزامی:\n\n"
        "دسته‌بندی | نام محصول | واحد | قیمت\n\n"
        "ستون اختیاری پنجم: وضعیت (فعال یا غیرفعال)\n\n"
        "همه سطرهای معتبر پردازش می‌شوند. اگر دسته‌بندی چند سطر پشت‌سرهم یکی است، "
        "می‌توانید فقط سطر اول آن را پر کنید و سطرهای بعدی را خالی بگذارید.\n"
        "محصول تکراری در همان دسته به‌روزرسانی می‌شود. برای لغو /cancel را بزنید."
    )
    return AWAIT_EXCEL_FILE


async def import_excel_receive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    if not document or not document.file_name.lower().endswith((".xlsx", ".xlsm")):
        await update.message.reply_text("لطفاً یک فایل اکسل معتبر (.xlsx یا .xlsm) بفرستید.")
        return AWAIT_EXCEL_FILE

    file = await context.bot.get_file(document.file_id)
    file_bytes = await file.download_as_bytearray()

    try:
        wb = load_workbook(
            io.BytesIO(bytes(file_bytes)), data_only=True, read_only=True
        )
        sheet = wb.active
    except Exception as exc:
        await update.message.reply_text(
            f"نتوانستم فایل را باز کنم. مطمئن شوید فایل سالم و با فرمت xlsx است. ({type(exc).__name__})"
        )
        return AWAIT_EXCEL_FILE

    header_row, columns = detect_header_map(sheet)
    if not columns:
        wb.close()
        await update.message.reply_text(
            "ستون‌های لازم پیدا نشد. حداقل چهار ستون دسته‌بندی، نام محصول، واحد و قیمت لازم است."
        )
        return AWAIT_EXCEL_FILE

    added = updated = skipped = 0
    errors = []
    last_category = None

    for excel_row_number, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue

        raw_category = cell_value(row, columns["category"])
        if raw_category is not None and str(raw_category).strip():
            last_category = str(raw_category).strip()
        category_name = last_category
        product_name = cell_value(row, columns["name"])
        unit = cell_value(row, columns["unit"])
        raw_price = cell_value(row, columns["price"])
        raw_active = cell_value(row, columns.get("active"))

        try:
            if not category_name:
                raise ValueError("دسته‌بندی خالی است و دسته قبلی هم وجود ندارد.")
            if product_name is None or not str(product_name).strip():
                raise ValueError("نام محصول خالی است.")
            if unit is None or not str(unit).strip():
                raise ValueError("واحد خالی است.")

            price = parse_price(raw_price)
            active = parse_active(raw_active)
            action, _ = db.upsert_product(
                category_name,
                str(product_name).strip(),
                str(unit).strip(),
                price,
                active,
            )
            if action == "added":
                added += 1
            else:
                updated += 1
        except Exception as exc:
            skipped += 1
            if len(errors) < 10:
                errors.append(f"سطر {excel_row_number}: {str(exc)}")

    wb.close()
    details = ""
    if errors:
        details = "\n\nجزئیات اولین خطاها:\n" + "\n".join(f"• {e}" for e in errors)
        if skipped > len(errors):
            details += f"\n• و {skipped - len(errors)} خطای دیگر"

    await update.message.reply_text(
        f"✅ ورود اکسل تمام شد:\n"
        f"➕ {added} محصول جدید اضافه شد\n"
        f"✏️ {updated} محصول به‌روزرسانی شد\n"
        f"⏭️ {skipped} سطر نامعتبر رد شد"
        f"{details}",
        reply_markup=kb.admin_menu_keyboard(),
    )
    return ConversationHandler.END


async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_only_guard(update):
        return
    cats = {c["id"]: c["name"] for c in db.list_categories()}
    products = db.list_products(active_only=False)
    if not products:
        await update.message.reply_text("هنوز محصولی ثبت نشده.")
        return

    wb = Workbook()
    sheet = wb.active
    sheet.title = "محصولات"
    sheet.append(["دسته‌بندی", "نام محصول", "واحد", "قیمت", "وضعیت"])
    for p in products:
        sheet.append([
            cats.get(p["category_id"], ""),
            p["name"],
            p["unit"],
            p["price"],
            "فعال" if p["is_active"] else "غیرفعال",
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    await update.message.reply_document(
        document=buffer,
        filename="محصولات.xlsx",
        caption="لیست فعلی همه محصولات؛ شامل محصولات فعال و غیرفعال",
    )


# ---------- مدیریت سفارش‌ها ----------

async def pending_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_only_guard(update):
        return
    orders = db.list_all_orders(status="pending")
    if not orders:
        await update.message.reply_text("سفارش در انتظاری وجود ندارد. 🎉")
        return
    for order in orders:
        await send_order_summary(update, context, order)


async def all_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_only_guard(update):
        return
    orders = db.list_all_orders(limit=20)
    if not orders:
        await update.message.reply_text("هنوز سفارشی ثبت نشده.")
        return
    for order in orders:
        await send_order_summary(update, context, order)


async def send_order_summary(update, context, order):
    user = db.get_user(order["user_id"])
    status_label = ORDER_STATUSES.get(order["status"], order["status"])
    text = (
        f"سفارش #{order['id']}\n"
        f"مشتری: {user['full_name'] if user else order['user_id']}\n"
        f"مبلغ: {order['total_price']:,} تومان\n"
        f"وضعیت: {status_label}\n"
        f"تاریخ: {order['created_at'][:16]}"
    )
    await update.message.reply_text(text, reply_markup=kb.order_status_keyboard(order["id"]))


async def set_order_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.edit_message_text("⛔ فقط برای ادمین.")
        return
    _, order_id, status_key = query.data.split(":")
    order_id = int(order_id)
    db.update_order_status(order_id, status_key)
    order = db.get_order(order_id)
    status_label = ORDER_STATUSES.get(status_key, status_key)
    await query.edit_message_text(f"وضعیت سفارش #{order_id} به «{status_label}» تغییر کرد.")

    try:
        await context.bot.send_message(
            order["user_id"],
            f"📦 وضعیت سفارش #{order_id} شما به‌روز شد:\n{status_label}",
        )
    except Exception:
        pass


async def review_receipt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not is_admin(query.from_user.id):
        await query.answer("فقط ادمین اجازه این کار را دارد.", show_alert=True)
        return
    action, receipt_id_text = query.data.split(":")
    receipt_id = int(receipt_id_text)
    receipt = db.get_payment_receipt(receipt_id)
    if not receipt:
        await query.edit_message_caption(caption="فیش پیدا نشد.")
        return
    if receipt["status"] != "pending_review":
        await query.answer("این فیش قبلاً بررسی شده است.", show_alert=True)
        return

    approved = action == "receiptapprove"
    status = "approved" if approved else "rejected"
    db.update_receipt_status(receipt_id, status, query.from_user.id)
    result = "✅ پرداخت تأیید شد" if approved else "❌ فیش رد شد"
    old_caption = query.message.caption or ""
    await query.edit_message_caption(caption=f"{old_caption}\n\n{result} — توسط ادمین")
    try:
        await context.bot.send_message(
            receipt["user_id"],
            f"{result}\nفیش #{receipt_id} مربوط به سفارش #{receipt['order_id']} بررسی شد."
            + ("" if approved else "\nلطفاً تصویر صحیح یا واضح‌تری ارسال کنید."),
        )
    except Exception:
        pass


def register_admin_handlers(app):
    app.add_handler(CommandHandler("admin", admin_panel))
    app.add_handler(MessageHandler(filters.Regex("^🔙 بازگشت به منوی کاربر$"), back_to_user_menu))
    app.add_handler(MessageHandler(filters.Regex("^🧰 مدیریت محصولات$"), manage_products_start))
    app.add_handler(MessageHandler(filters.Regex("^📦 سفارش‌های در انتظار$"), pending_orders))
    app.add_handler(MessageHandler(filters.Regex("^📊 همه سفارش‌ها$"), all_orders))

    # دکمه‌های مستقیم مدیریت محصول
    app.add_handler(CallbackQueryHandler(manage_products_page, pattern=r"^managepage:\d+$"))
    app.add_handler(CallbackQueryHandler(manage_product_open, pattern=r"^manageprod:\d+:\d+$"))
    app.add_handler(CallbackQueryHandler(manage_product_toggle, pattern=r"^prodtoggle:\d+:\d+$"))
    app.add_handler(CallbackQueryHandler(manage_product_delete_ask, pattern=r"^proddeleteask:\d+:\d+$"))
    app.add_handler(CallbackQueryHandler(manage_product_delete, pattern=r"^proddelete:\d+:\d+$"))
    app.add_handler(CallbackQueryHandler(manage_close, pattern=r"^manageclose$"))
    app.add_handler(CallbackQueryHandler(manage_noop, pattern=r"^managenoop$"))

    app.add_handler(CallbackQueryHandler(edit_price_pick_category, pattern=r"^editcat:\d+$"))
    app.add_handler(CallbackQueryHandler(set_order_status, pattern=r"^setstatus:\d+:\w+$"))
    app.add_handler(CallbackQueryHandler(review_receipt, pattern=r"^receipt(?:approve|reject):\d+$"))

    app.add_handler(MessageHandler(filters.Regex("^📤 خروجی اکسل$"), export_excel))

    import_excel_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📥 وارد کردن از اکسل$"), import_excel_start)],
        states={
            AWAIT_EXCEL_FILE: [MessageHandler(filters.Document.ALL, import_excel_receive)],
        },
        fallbacks=[CommandHandler("cancel", cancel_admin_operation)],
    )
    app.add_handler(import_excel_conv)

    add_product_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^➕ افزودن محصول$"), add_product_start)],
        states={
            ADD_CAT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_product_category)],
            ADD_PROD_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_product_name)],
            ADD_PROD_UNIT: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_product_unit)],
            ADD_PROD_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_product_price)],
        },
        fallbacks=[CommandHandler("cancel", cancel_admin_operation)],
    )
    app.add_handler(add_product_conv)

    edit_product_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(manage_product_edit_start, pattern=r"^prodedit:\d+:\d+$")],
        states={
            MANAGE_EDIT_CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, manage_edit_category)],
            MANAGE_EDIT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, manage_edit_name)],
            MANAGE_EDIT_UNIT: [MessageHandler(filters.TEXT & ~filters.COMMAND, manage_edit_unit)],
            MANAGE_EDIT_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manage_edit_price)],
        },
        fallbacks=[CommandHandler("cancel", cancel_admin_operation)],
    )
    app.add_handler(edit_product_conv)

    # نگه‌داشتن مسیر قدیمی تغییر قیمت برای کاربرانی که دکمه قدیمی هنوز در تلگرامشان دیده می‌شود.
    edit_price_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^✏️ تغییر قیمت$"), edit_price_start),
            CallbackQueryHandler(edit_price_pick_product, pattern=r"^editprod:\d+$"),
        ],
        states={
            EDIT_PRICE_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_price_value)],
        },
        fallbacks=[CommandHandler("cancel", cancel_admin_operation)],
    )
    app.add_handler(edit_price_conv)
